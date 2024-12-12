import boto3
import io

from pprint import pprint
from random import randint
from time import sleep
from openpyxl import load_workbook
from django.conf import settings
from urllib.request import urlopen
from saleor.product.models import ProductVariant
from saleor.warehouse.models import Stock
from ...celeryconf import app
from ..email_handler import send_email_report

@app.task
def import_inventory_data(manual=False):
    pprint("*********************************************")
    pprint("************* vox_update_stock **************")
    pprint("*********************************************")

    email_report_message = ""
    errors = ""
    updated_count = 0
    error_count = 0

    # Try to prevent the same job running multiple times.  The site is load banaced
    # so the job is running multiple times causing duplicate inventory updates. When
    # a job start and completes the status will be update but there is a race condition
    # when the status doesn't update quickly enough.
    if manual:
        delay_seconds = 1
    else:
        delay_seconds = randint(10,10800)

    pprint("Delaying Import by "+str(delay_seconds)+" seconds")
    sleep(delay_seconds)

    try:
        s3_client = boto3.client(
            'workdocs',
            aws_access_key_id=settings.AWS_DATA_ACCESS_KEY_ID,
            aws_secret_access_key=settings.AWS_DATA_SECRET_ACCESS_KEY,
            region_name=settings.AWS_WORKDOCS_REGION
        )

        get_document_response = s3_client.get_document(
            DocumentId=settings.AWS_WORKDOCS_INVENTORY_DOCUMENT_ID,
            IncludeCustomMetadata=True,
        )
        version_id = get_document_response["Metadata"]["LatestVersionMetadata"]["Id"]

        document = s3_client.get_document_version(
            DocumentId=settings.AWS_WORKDOCS_INVENTORY_DOCUMENT_ID,
            VersionId=version_id,
            IncludeCustomMetadata=True,
            Fields='Source'
        )

        if "CustomMetadata" in document:
            if "ImportStatus" in document['CustomMetadata']:
                if document['CustomMetadata']['ImportStatus']:
                    if document['CustomMetadata']['ImportStatus'] == "SUCCESS" or document['CustomMetadata']['ImportStatus'] == "IMPORTING":
                        email_report_message = "The Inventory file has already been imported"
                        return False, email_report_message

        # update the import status
        s3_client.create_custom_metadata(
            ResourceId=settings.AWS_WORKDOCS_INVENTORY_DOCUMENT_ID,
            VersionId=version_id,
            CustomMetadata={
                'ImportStatus': 'IMPORTING'
            }
        )

        f = urlopen(document['Metadata']['Source']['ORIGINAL'])
        myfile = f.read()
        wb = load_workbook(io.BytesIO(myfile))
        ws = wb.active

        for row in ws.iter_rows(min_row=4):
            sku = row[0].value
            qty = row[2].value
            if sku:
                if qty is not None and isinstance(qty, int) and int(qty) > 0:
                    try:
                        variant = ProductVariant.objects.using("replica").get(sku=sku)
                    except ProductVariant.DoesNotExist:
                        pprint("Inventory Update Sku Not found: " + str(sku))
                        errors += "SKU Not Found: "+str(sku)+"\n"
                        error_count+=1
                        variant = None
                    if variant:
                        stock = Stock.objects.get(product_variant=variant, warehouse_id="171cb0c0-90a8-48d1-bfe3-3201ddabce8d")
                        if stock:
                            stock.quantity = stock.quantity + qty
                            stock.save()

                            updated_count+=1
                            pprint("Inventory Update: " + str(sku) + ":" + str(qty))

            else:
                pprint("Inventory Update SKU Not Found: "+str(sku))
                if sku is not None:
                    errors += "SKU Not Found: "+str(sku)+"\n"
                    error_count+=1

        # update the import status
        s3_client.create_custom_metadata(
            ResourceId=settings.AWS_WORKDOCS_INVENTORY_DOCUMENT_ID,
            VersionId=version_id,
            CustomMetadata={
                'ImportStatus': 'SUCCESS'
            }
        )

        email_report_message = "The inventory has been imported successfully"

        return True, email_report_message

    except Exception as e:
        email_report_message = "An error occurred while importing the inventory file: "+str(e)
        return False, email_report_message
    finally:
        send_email_report("Inventory Import Report", email_report_message, errors, 0, updated_count, error_count)
